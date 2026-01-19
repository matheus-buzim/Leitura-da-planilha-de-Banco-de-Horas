from pathlib import Path
from openpyxl import load_workbook

import sys
import os

def read_cell(folder_path: str, cell: str) -> list[dict]:

    results = []

    for file in Path(folder_path).glob("*.xlsx"):
        planilha = load_workbook(file, data_only=True)

        for abas in planilha.sheetnames:
            aba = planilha[abas]
            value = aba[cell].value

            results.append({
                "file": file.name,
                "sheet": abas,
                "cell": cell,
                "value": value,
            })

    return results

def main():

    base_dir = Path(sys.executable).resolve().parent
    folder = os.path.join(base_dir, "input")
    celula = "L46" # Aqui eu coloco a célula que contém o total de horas, é sempre a mesma célula em cada aba

    data = read_cell(folder_path=folder, cell=celula)

    if not data:
        print("Não foi encontrado nenhum arquivo.")
        return

    arquivo_atual = None

    for item in data:

        #A planilha tem abas antigas ocultas, como ela já foi disponibilizada assim,
        #para não ter o retrabalho de reenvio, eu simplesmente ocultei as antigas.
        sheet = item["sheet"].strip().lower()

        if "25" in sheet:
            continue

        # Existe na planilha da empresa, uma aba que contém as configurações de horas, sendo assim, ocultei da lista.
        if item["sheet"] == "Não Apagar":
            continue

        if item["file"] != arquivo_atual: #Aqui eu apenas faço uma manobra para separar a lista a cada arquivo lido.
            if arquivo_atual is not None:
                print("\n" + "-" * 43 + "\n")

            arquivo_atual = item["file"]
            print(arquivo_atual + "\n")


        print(f"   {item['sheet']} | {item['value']} hrs")

    input("\n\nAperte ENTER para fechar a tela.")

if __name__ == "__main__":
    main()